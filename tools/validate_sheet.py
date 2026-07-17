# -*- coding: utf-8 -*-
"""Validate the product workbook BEFORE gen_regions.py runs (CI gate).

Usage: python3 tools/validate_sheet.py <path-to-xlsx>
Exits non-zero with a clear message if the sheet would break generation.
Never touches the site — this is the fail-loud guard so a bad edit in the
Google Sheet can never take the live site down.
"""

import sys, os

import openpyxl

REQUIRED_SHEET = "产品总表"
REQUIRED_COLS = [
    "板块region",
    "分组group",
    "产品名称name",
    "TourCode",
    "天数days",
    "出发城市dep",
    "抵达城市arr",
    "成人价adult",
    "儿童占床child_with",
    "儿童不占床child_no",
    "单房差single",
    "其他费other",
    "banner图img",
    "行程安排itinerary",
    "参团须知notes",
]
MANAGED = {"china", "asia", "island", "america", "europe", "other"}
PRICE_COLS = ["成人价adult", "儿童占床child_with", "儿童不占床child_no", "单房差single"]


def fail(msg) -> "NoReturn":
    print("[FAIL] " + msg)
    sys.exit(1)


from typing import NoReturn  # noqa: E402


def main():
    path = sys.argv[1]
    if not os.path.exists(path):
        fail("文件不存在: " + path)
    if os.path.getsize(path) < 1000:
        fail(
            "下载的文件太小，多半不是 Excel。请检查 Google Sheet 是否已开启"
            "「知道链接的任何人可查看」共享。"
        )
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        fail("无法打开 Excel（Google Sheet 未共享时会下载到登录页 HTML）: %s" % e)
    if REQUIRED_SHEET not in wb.sheetnames:
        fail("缺少工作表「%s」。现有: %s" % (REQUIRED_SHEET, wb.sheetnames))
    ws = wb[REQUIRED_SHEET]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in REQUIRED_COLS if c not in hdr]
    if missing:
        fail("表头缺少这些列（列名必须完全一致）: %s" % missing)

    idx = {c: hdr.index(c) for c in REQUIRED_COLS}
    rows = 0
    regions = {}
    problems = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(len(hdr))]
        name = vals[idx["产品名称name"]]
        if not name:
            continue
        rows += 1
        region = str(vals[idx["板块region"]] or "").strip()
        regions[region] = regions.get(region, 0) + 1
        if region and region not in MANAGED:
            problems.append(
                "第%d行: 板块「%s」不在受管板块 %s 中（该行会被忽略）"
                % (r, region, sorted(MANAGED))
            )
        for pc in PRICE_COLS:
            v = vals[idx[pc]]
            if v in (None, ""):
                continue
            try:
                float(str(v))
            except (TypeError, ValueError):
                problems.append(
                    "第%d行「%s」: %s 不是数字（'%s'）— 会显示为「待确认」"
                    % (r, name, pc, v)
                )

    if rows == 0:
        fail("没有任何数据行（产品名称列全为空）。")

    for p in problems:
        print("[WARN] " + p)
    print("[OK] %d 条产品, 各板块: %s" % (rows, regions))


if __name__ == "__main__":
    main()
