# -*- coding: utf-8 -*-
"""
一键生成: 读 [产品总表.xlsx] -> 重建 region-plans.js 中 china/asia/island/america/europe/other 六个块
保留 australia/nz/cruise (不受总表管理, 不动)
用法:
  1. 用 Excel 编辑 D:/Hermes Agent/etrips-assets/产品总表.xlsx (改产品/价格/行程/须知)
  2. cd C:/Users/linda/etrips-site
  3. python tools/gen_regions.py
  4. git add -A && git commit && git push
数据源/产物路径说明:
  SRC = 产品总表.xlsx (默认从 D:\Hermes Agent\etrips-assets\ 读取; 也可用环境变量 ET_PRODUCT_XLSX 覆盖)
  JS  = 本仓库 assets/js/region-plans.js
"""

import openpyxl, re, io, os

SRC = os.environ.get("ET_PRODUCT_XLSX", r"D:/Hermes Agent/etrips-assets/产品总表.xlsx")
JS = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets",
    "js",
    "region-plans.js",
)
MANAGED = ["china", "asia", "island", "america", "europe", "other"]


def norm(s):
    return re.sub(r"\s+", "", str(s)).strip()


def slug(name):
    return re.sub(r"[^一-龥a-zA-Z0-9]", "", str(name))[:18]


def fmt_price(v):
    try:
        f = float(v)
        return "待确认" if f == 0 else (str(int(f)) if f == int(f) else str(f))
    except:
        return "待确认"


def days_from(d, name):
    if d:
        return str(d) + "天"
    m = re.search(r"(\d+)\s*日", str(name))
    return (m.group(1) + "天") if m else "待确认"


# --- 读总表 ---
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["产品总表"]
rows = []
hdr = [c.value for c in ws[1]]
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, len(hdr) + 1)]
    if not vals[2]:
        continue
    rows.append(dict(zip(hdr, vals)))

# route id 去重
seen = {}


def rid(region, name, code):
    base = (
        region[:2]
        + "-"
        + (slug(name) if not code else re.sub(r"[^a-z0-9]", "", str(code).lower()))
    )
    r = base
    i = 1
    while r in seen:
        i += 1
        r = base + str(i)
    seen[r] = True
    return r


IMG_POOL = {
    "china": ["china.jpg", "cn-westlake.jpg", "cn-greatwall.jpg"],
    "asia": ["asia.jpg", "cn-westlake.jpg", "japan.jpg"],
    "island": ["island.jpg", "bali.jpg", "fiji.jpg"],
    "america": ["america.jpg", "canada.jpg", "usa.jpg"],
    "europe": ["europe.jpg", "paris.jpg", "greece.jpg"],
    "other": ["other.jpg", "custom.jpg"],
}


def price_table(adult, cno, cw, single, other):
    items = [
        ("成人报名费 Tour Fee", "Adult", adult),
        ("儿童不占床 Child no bed", "Child no bed", cno),
        ("儿童占床 Child with bed", "Child with bed", cw),
    ]
    if single not in ("", None):
        items.append(("单房差 Single suppl.", "Single", single))
    if other not in ("", None):
        items.append(("综合服务费 Service fee", "Service", other))
    body = "".join(
        '<tr><td class="item">'
        + z
        + '<span class="en">'
        + en
        + '</span></td><td class="price">'
        + fmt_price(p)
        + "</td></tr>"
        for z, en, p in items
    )
    return (
        '<table class="rp-pricetable"><thead><tr><th>项目</th><th style="text-align:right">价格 (AUD)</th></tr></thead><tbody>'
        + body
        + "</tbody></table>"
    )


def pane(region, item, idx):
    name = item["产品名称name"] or ""
    days = days_from(item.get("天数days"), name)
    adult = item.get("成人价adult")
    cno = item.get("儿童不占床child_no")
    cw = item.get("儿童占床child_with")
    single = item.get("单房差single")
    other = item.get("其他费other")
    pa = (
        float(adult)
        if (adult not in ("", None) and str(adult).replace(".", "").isdigit())
        else 0
    )
    pc = (
        float(cno)
        if (cno not in ("", None) and str(cno).replace(".", "").isdigit())
        else 0
    )
    ridv = item[
        "_rid"
    ]  # rid 只算一次(在 build_block), 避免去重后缀导致导航/详情卡 data-route 不一致
    img = "assets/img/destinations/" + (
        item.get("banner图img") or IMG_POOL[region][idx % len(IMG_POOL[region])]
    )
    pax = ("A$ " + str(int(pa))) if pa else "待确认"
    pcx = ("A$ " + str(int(pc))) if pc else "待确认"
    badge = {"island": "海岛度假", "america": "美加专线"}.get(region, "纯玩无购物")
    iti = item.get("行程安排itinerary") or "【行程安排】\n请在此处粘贴行程安排内容..."
    notes = item.get("参团须知notes") or "【参团须知】\n请在此处粘贴参团须知内容..."
    return (
        '    <div class="rp-route-pane" data-route="'
        + ridv
        + '" data-p-adult="'
        + str(int(pa) if pa else 0)
        + '" data-p-child="'
        + str(int(pc) if pc else 0)
        + '" data-p-infant="0">\n'
        '      <div class="rp-detail-hero" style="background-image:url(\''
        + img
        + "')\">\n"
        '        <span class="rp-badge">' + badge + "</span>\n"
        '        <div class="rp-detail-hero-in"><h3>' + name + "</h3>\n"
        '          <div class="rp-price-row">\n'
        '            <span class="rp-price-item"><b>大人</b> ' + pax + "</span>\n"
        '            <span class="rp-price-item"><b>儿童</b> ' + pcx + "</span>\n"
        '            <span class="rp-price-item"><b>婴儿</b> 待确认</span>\n'
        "          </div>\n"
        "        </div>\n"
        "      </div>\n"
        '      <div class="rp-tabs">\n'
        '        <div class="rp-tab active" data-tab="price">日期和价格</div>\n'
        '        <div class="rp-tab" data-tab="itinerary">行程安排</div>\n'
        '        <div class="rp-tab" data-tab="notes">参团须知</div>\n'
        '        <div class="rp-tab" data-tab="brochure">彩页下载(澳洲)</div>\n'
        "      </div>\n"
        '      <div class="rp-tab-panel active" data-tab="price">\n'
        '        <div class="rp-summary"><div><b>行程天数</b>'
        + days
        + "</div><div><b>抵达城市</b>"
        + str(item.get("抵达城市arr") or "待确认")
        + "</div><div><b>离开城市</b>"
        + str(item.get("出发城市dep") or "待确认")
        + "</div></div>\n"
        + price_table(adult, cno, cw, single, other)
        + '        <p style="margin-top:12px;font-size:12px;color:#8a97a6">详情与班期以客服查询为准。</p>\n'
        "      </div>\n"
        '      <div class="rp-tab-panel" data-tab="itinerary"><pre style="white-space:pre-wrap;font-size:14px;line-height:1.7;color:#334">\n'
        + iti
        + "\n</pre></div>\n"
        '      <div class="rp-tab-panel" data-tab="notes"><pre style="white-space:pre-wrap;font-size:14px;line-height:1.7;color:#334">\n'
        + notes
        + "\n</pre></div>\n"
        '      <div class="rp-tab-panel" data-tab="brochure"><p>彩页下载：请咨询客服获取 PDF 彩页。</p></div>\n'
        "    </div>"
    )


from collections import OrderedDict

BANNER_TITLE = {
    "china": ("中国 · 分区行程规划", "China Tours"),
    "asia": ("亚洲 · 分区行程规划", "Asia Tours"),
    "island": ("海岛假日 · 度假天堂", "Island Holidays"),
    "america": ("美国 · 加拿大 · 南美", "America & Canada"),
    "europe": ("欧洲 · 经典环游", "Europe Tours"),
    "other": ("其他 · 更多目的地", "Other Destinations"),
}
BANNER_IMG = {
    "china": ["china.jpg", "cn-westlake.jpg"],
    "asia": ["asia.jpg", "cn-westlake.jpg"],
    "island": ["island.jpg", "bali.jpg", "fiji.jpg"],
    "america": ["america.jpg", "canada.jpg", "usa.jpg"],
    "europe": ["europe.jpg", "paris.jpg", "greece.jpg"],
    "other": ["other.jpg", "custom.jpg"],
}
CREDIT = "Banner 图片：© Wikimedia (Public Domain / CC)."


def build_block(region, items):
    tree = OrderedDict()
    for it in items:
        g = it.get("分组group") or "其他"
        tree.setdefault(g, []).append(it)
    nav = ""
    panes = ""
    idx = 0
    for g, its in tree.items():
        for x in its:
            x["_rid"] = rid(
                region, (x.get("产品名称name") or ""), x.get("TourCode") or ""
            )
        routes = "".join(
            '        <div class="rp-route" data-route="'
            + x["_rid"]
            + '">'
            + (x.get("产品名称name") or "")
            + "</div>\n"
            for x in its
        )
        nav += (
            '      <div class="rp-group" data-group="'
            + g
            + '">\n        <div class="rp-group-title">'
            + g
            + ' <span class="rp-arrow">▶</span></div>\n        <div class="rp-group-list">\n'
            + routes
            + "        </div>\n      </div>\n"
        )
        for x in its:
            panes += pane(region, x, idx)
            idx += 1
    t, sub = BANNER_TITLE[region]
    slides = "".join(
        '    <div class="rp-slide'
        + (
            (" active" if i == 0 else "")
            + '" style="background-image:url(\'assets/img/destinations/'
            + b
            + "')\"></div>\n"
        )
        for i, b in enumerate(BANNER_IMG[region])
    )
    return (
        "  window.REGION_PLANS." + region + " = `\n"
        '<div class="rp-banner">\n  <div class="rp-slides">\n' + slides + "  </div>\n"
        '  <div class="rp-banner-in">\n    <h1>'
        + t
        + '</h1>\n    <div class="rp-sub">'
        + sub
        + "</div>\n"
        '    <p class="rp-desc">点击左侧区域，查看各地核心行程。详情与班期以客服查询为准。</p>\n  </div>\n</div>\n'
        '<div class="rp-layout">\n  <nav class="rp-nav2" aria-label="'
        + region
        + ' 目的地">\n'
        + nav
        + "  </nav>\n"
        '  <div class="rp-detail-area">\n' + panes + "  </div>\n</div>\n\n"
        '  <p class="rp-banner-credit" style="color:#999;font-size:12px;margin:14px 0 0 0;">'
        + CREDIT
        + "</p>`;"
    )


# 按板块分组
by_region = OrderedDict()
for it in rows:
    rg = it.get("板块region") or ""
    if rg in MANAGED:
        by_region.setdefault(rg, []).append(it)

blocks = {rg: build_block(rg, its) for rg, its in by_region.items()}

# --- 替换 region-plans.js 中受管块 ---
# 字符串切分 + 反引号配平: 找每个 window.REGION_PLANS.X = ` 起点, 配平到结束 `
s = io.open(JS, encoding="utf-8").read()
for k in MANAGED:
    if k not in blocks:
        continue
    marker = "window.REGION_PLANS." + k + " = `"
    st = s.index(marker)
    # 从首个 ` 之后配平
    i = s.index("`", st) + 1
    depth = 1
    while i < len(s) and depth > 0:
        if s[i] == "`":
            depth -= 1
        i += 1
    end = i  # 结束 ` 之后
    s = s[:st] + blocks[k] + s[end:]
io.open(JS, "w", encoding="utf-8").write(s)
# 验证
import subprocess

for k in MANAGED:
    assert ("window.REGION_PLANS." + k + " = `") in s, k
print(
    "region-plans.js 已重建. 受管板块:", {k: len(by_region.get(k, [])) for k in MANAGED}
)
r = subprocess.run(
    [
        "node",
        "-e",
        "new Function(require('fs').readFileSync(process.argv[1],'utf8'))",
        JS,
    ],
    capture_output=True,
    text=True,
)
print("node 语法:", "OK" if r.returncode == 0 else r.stderr[:80])
