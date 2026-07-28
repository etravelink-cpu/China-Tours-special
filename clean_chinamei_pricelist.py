# -*- coding: utf-8 -*-
"""
清洗 中国美 price list 2026.xlsx -> 中国美_price_cleaned.xlsx
结构差异:
  中国/亚洲 sheet: 团名行 + 性质行(超值特价/纯玩无购物/含机票) 配对 -> 合并, 类目由性质行定
  美加/海岛/欧洲 sheet: 团名行(含天数) + 价格变体行(如"1月早鸟AUD388"/"纽约+多伦多"/"黄石峡谷") 配对 -> 合并, 类目按sheet默认(纯玩无购物团, 含"包机票"->含机票)
  模板组: "2026XX定制团" + "非常规团 9999" -> 排除(非真实上架)
规则(据用户):
  中国/亚洲/美加/欧洲 -> 各自地区板块, 类目按性质行或sheet默认
  海岛/全球签证 -> 统一归 "签证·其他" (放"其他"板块)
"""
import openpyxl, re, os

SRC = r"D:/Hermes Agent/对话备份/.hermes/desktop-attachments/中国美 price list 2026.xlsx"
OUT = r"D:/Hermes Agent/etrips-universal-db/中国美_price_cleaned.xlsx"

PROP_MAP = {
    "超值特价": "超值特惠团", "超值特惠": "超值特惠团", "超值": "超值特惠团",
    "纯玩无购物": "纯玩无购物团", "纯玩": "纯玩无购物团",
    "含机票": "含机票特别订制团", "包机票": "含机票特别订制团",
}
SHEET_CFG = {
    "中国":   ("中国", "by_prop"),
    "亚洲":   ("亚洲", "by_prop"),
    "美加":   ("美加", "pure_play_default"),
    "欧洲":   ("欧洲", "pure_play_default"),
    "海岛":   ("海岛", "visa_other"),     # 放"其他"
    "全球签证": ("其他", "visa_other"),    # 放"其他"
}
OUT_COLS = ["Supplier","Supplier_Product_Code","Internal_Product_Code","Product_Category",
            "Region","Country","Product_Name_CN","Product_Name_EN","Duration_Days",
            "Adult_Price_AUD","Child_Price_AUD","Child_With_Bed_AUD","Child_No_Bed_AUD",
            "Infant_Price_AUD","Single_Supplement_AUD","Airport_Transfer_AUD",
            "Mandatory_Tips_AUD","Remarks","Status"]

def to_num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return v
    s=str(v).strip().replace("AUD","").replace("$","").replace(",","").strip()
    if s=="" or s.upper()=="AUD": return None
    try: return float(s)
    except: return None

def has_days(nm):
    return bool(re.search(r"\d+\s*(日|天|晚|夜)", str(nm)))

def is_real_product(nm):
    nm=str(nm).strip()
    if has_days(nm): return True
    if re.search(r"(之旅|全景|游轮|邮轮|定制团|订制团|特别团|狂欢团|报名团|深度游|观光|经典.*游|全景游|全景之旅|精华.*游|之旅|考察|探险)", nm): return True
    return False

def is_template(nm, adult):
    """模板占位组: 标题'2026XX定制团' + 非常规团 9999"""
    return ("定制团" in str(nm)) or (adult is not None and adult >= 9999) or ("非常规团" in str(nm))

def is_prop_or_variant(nm, adult):
    """性质行(超值特价/纯玩) 或 价格变体行(无天数, 但带价格数字 或 行程变体描述)"""
    nm=str(nm).strip()
    if has_days(nm): return False
    if re.search(r"(之旅|全景|游轮|邮轮|定制团|订制团|特别团|狂欢团|报名团|深度游|观光|经典.*游|全景游|全景之旅|精华.*游|之旅|考察|探险)", nm): return False
    # 性质词 或 含价格数字(变体行有真实价) 或 行程变体关键词
    if any(k in nm for k in PROP_MAP): return True
    if adult is not None and adult > 0: return True
    if re.search(r"(早鸟|纯玩|无购物|黄石|尼亚加拉|多伦多|经典|跨岛|五星|升级|不含|含)", nm): return True
    return False

def prop_cat(nm):
    nm=str(nm).strip()
    for k,v in PROP_MAP.items():
        if k in nm: return v
    return None

wb=openpyxl.load_workbook(SRC, data_only=True)
out_wb=openpyxl.Workbook(); out_ws=out_wb.active
out_ws.title="Master_Product_Database"; out_ws.append(OUT_COLS)

total=0; placeholder=0; skipped=0
for sheet,(wc1,cat_mode) in SHEET_CFG.items():
    if sheet not in wb.sheetnames: continue
    ws=wb[sheet]; rows=list(ws.iter_rows(values_only=True))
    pending=None
    for r in rows[1:]:
        nm=r[2] if len(r)>2 else None
        if nm is None or str(nm).strip()=="": continue
        nm_str=str(nm).strip()
        adult=to_num(r[3]) if len(r)>3 else None
        # 合并行(性质/变体) -> 补价格到 pending
        if is_prop_or_variant(nm_str, adult):
            if pending is not None:
                pr=list(pending)
                for src,dst in [(3,"Adult_Price_AUD"),(4,"Child_No_Bed_AUD"),
                                (6,"Child_With_Bed_AUD"),(7,"Single_Supplement_AUD"),
                                (8,"Mandatory_Tips_AUD"),(9,"Airport_Transfer_AUD")]:
                    if src<len(r):
                        nv=to_num(r[src])
                        if nv is not None: pr[OUT_COLS.index(dst)]=nv
                if cat_mode=="by_prop":
                    pc=prop_cat(nm_str)
                    if pc: pr[OUT_COLS.index("Product_Category")]=pc
                elif "包机票" in nm_str or "含机票" in nm_str:
                    pr[OUT_COLS.index("Product_Category")]="含机票特别订制团"
                out_ws.append(pr); total+=1; pending=None
            continue
        if is_template(nm_str, adult):
            placeholder+=1
            if pending is not None: pending=None  # 模板标题也丢
            continue
        if is_real_product(nm_str):
            adult2=to_num(r[3]) if len(r)>3 else None
            if adult2 is not None and adult2>=9999:
                placeholder+=1; continue
            row=[None]*len(OUT_COLS)
            row[OUT_COLS.index("Supplier")]="中国美"
            row[OUT_COLS.index("Product_Name_CN")]=nm_str
            row[OUT_COLS.index("Region")]=wc1
            row[OUT_COLS.index("Remarks")]=(r[12] if len(r)>12 else None)
            row[OUT_COLS.index("Status")]="Active"
            if cat_mode=="by_prop":
                row[OUT_COLS.index("Product_Category")]="纯玩无购物团"
            elif cat_mode=="pure_play_default":
                row[OUT_COLS.index("Product_Category")]="纯玩无购物团"
            elif cat_mode=="visa_other":
                row[OUT_COLS.index("Product_Category")]="签证·其他"
            if "包机票" in nm_str or "含机票" in nm_str:
                row[OUT_COLS.index("Product_Category")]="含机票特别订制团"
            pending=row
        else:
            skipped+=1
    if pending is not None:
        out_ws.append(pending); total+=1

out_wb.save(OUT)
print(f"清洗完成 -> {OUT}")
print(f"  产品总数: {total}")
print(f"  排除模板/占位(9999定制团): {placeholder}")
print(f"  跳过解释行: {skipped}")
